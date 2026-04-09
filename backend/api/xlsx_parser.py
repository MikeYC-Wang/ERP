"""XLSX parser for bulk product import.

Extracts rows from supplier order-sheet xlsx files with varying layouts.
Handles multi-section sheets (sections separated by a blank row + new header).
"""
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl


HEADER_KEYWORDS = {
    'sku': ['貨號', 'sku', '品號', '編號', '商品編號'],
    'name': ['品名', '商品名稱', '商品'],
    'sub_name': ['名稱'],
    'series': ['系列', '分類', '品牌'],
    'size': ['規格', '重量', '容量'],
    'unit': ['單位'],
    'unit_price': ['進貨價', '單價', '零售', '建議售價'],
    'box_price': ['一箱', '一組', '箱價', '整箱', '一盒'],
    'barcode': ['國際條碼', '條碼', 'barcode', 'ean'],
    'barcode_box': ['國際條碼(箱)', '條碼(箱)'],
}

# Header detection: any row containing >= 3 of these tokens qualifies.
HEADER_HINT_TOKENS = [
    '貨號', 'sku', '品號', '名稱', '品名', '規格', '單位',
    '箱', '價', '條碼', '系列', '進貨',
]


def _norm(v: Any) -> str:
    if v is None:
        return ''
    return str(v).strip().lower()


def _is_header_row(cells: list[Any]) -> bool:
    text_cells = [_norm(c) for c in cells if c is not None and str(c).strip()]
    if len(text_cells) < 3:
        return False
    hits = 0
    joined = ' '.join(text_cells)
    for tok in HEADER_HINT_TOKENS:
        if tok.lower() in joined:
            hits += 1
    return hits >= 3


def _map_headers(cells: list[Any]) -> dict[str, int]:
    """Return {canonical_key: column_index}. Later columns override earlier on tie."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(cells):
        t = _norm(cell)
        if not t:
            continue
        # barcode_box (check first — more specific than barcode)
        if '條碼' in t and ('箱' in t or '盒' in t):
            mapping['barcode_box'] = idx
            continue
        for key, kws in HEADER_KEYWORDS.items():
            if key == 'barcode_box':
                continue
            for kw in kws:
                if kw.lower() in t:
                    # For ambiguous tokens, prefer first match; but unit_price/box_price
                    # should prefer their rightmost match.
                    if key in ('unit_price', 'box_price') or key not in mapping:
                        mapping[key] = idx
                    break
    return mapping


_PACK_QTY_RE = re.compile(r'(\d+)\s*[^\d/]*[/／]?\s*[箱盒包]')


def _parse_unit_cell(val: Any) -> tuple[int, str]:
    """'24罐/箱' → (24, '罐'). '個' → (1, '個'). '包' → (1, '包')."""
    if val is None:
        return 1, '個'
    s = str(val).strip()
    if not s:
        return 1, '個'
    m = _PACK_QTY_RE.search(s)
    if m:
        qty = int(m.group(1))
        # base unit is the char(s) between the digits and '/箱'
        inner = s[m.start():m.end()]
        # pull out base unit: chars after digits, before '/箱' etc.
        sub = re.sub(r'^\d+\s*', '', inner)
        sub = re.sub(r'[/／]?\s*[箱盒包].*$', '', sub)
        base = sub.strip() or '個'
        return qty, base
    # No pack format — the whole string is the base unit
    return 1, s


def _to_decimal(val: Any) -> Decimal | None:
    if val is None or val == '':
        return None
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def _to_int(val: Any) -> int | None:
    d = _to_decimal(val)
    if d is None:
        return None
    try:
        return int(d)
    except (InvalidOperation, ValueError):
        return None


def parse_workbook(file_obj, all_sheets: bool = False) -> tuple[list[dict], list[str]]:
    """Parse an xlsx file object and return (rows, warnings)."""
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    rows: list[dict] = []
    warnings: list[str] = []
    seen_skus: set[str] = set()

    sheets = wb.sheetnames if all_sheets else wb.sheetnames[:1]

    for sheet_name in sheets:
        ws = wb[sheet_name]
        # Materialize rows (read_only ws is a generator)
        sheet_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        max_cols = max((len(r) for r in sheet_rows), default=0)

        i = 0
        n = len(sheet_rows)
        skipped_ranges: list[tuple[int, int]] = []
        while i < n:
            row = sheet_rows[i]
            if _is_header_row(row):
                headers = _map_headers(row)
                i += 1
                section_start = i
                no_sku_from = None
                while i < n:
                    data_row = sheet_rows[i]
                    # Blank row → end of section
                    if not any(c not in (None, '') for c in data_row):
                        i += 1
                        break
                    # Could be a new header row → stop section, let outer loop handle
                    if _is_header_row(data_row):
                        break

                    parsed = _row_to_product(data_row, headers, sheet_name, i + 1, warnings, seen_skus)
                    if parsed is None:
                        if no_sku_from is None:
                            no_sku_from = i + 1
                    else:
                        if no_sku_from is not None:
                            warnings.append(
                                f"Sheet '{sheet_name}' rows {no_sku_from}-{i} had no SKU, skipped"
                            )
                            no_sku_from = None
                        rows.append(parsed)
                    i += 1
                if no_sku_from is not None:
                    warnings.append(
                        f"Sheet '{sheet_name}' rows {no_sku_from}-{i} had no SKU, skipped"
                    )
            else:
                i += 1

    return rows, warnings


def _row_to_product(
    cells: list[Any],
    headers: dict[str, int],
    sheet_name: str,
    row_number: int,
    warnings: list[str],
    seen_skus: set[str],
) -> dict | None:
    def get(key: str) -> Any:
        idx = headers.get(key)
        if idx is None or idx >= len(cells):
            return None
        return cells[idx]

    sku_raw = get('sku')
    if sku_raw is None or str(sku_raw).strip() == '':
        return None
    sku = str(sku_raw).strip()
    if sku.lower() in ('貨號', 'sku'):  # stray header
        return None
    if sku in seen_skus:
        warnings.append(f"Sheet '{sheet_name}' row {row_number}: duplicate SKU '{sku}' skipped")
        return None
    seen_skus.add(sku)

    # Name: combine series/name/sub_name/size
    parts = []
    series = get('series')
    name = get('name')
    sub = get('sub_name')
    size = get('size')
    if name and str(name).strip():
        parts.append(str(name).strip())
    if sub and str(sub).strip() and (not name or str(sub).strip() != str(name).strip()):
        parts.append(str(sub).strip())
    if size and str(size).strip() and str(size).strip() != '-':
        parts.append(str(size).strip())
    full_name = ' '.join(parts) if parts else sku
    if series and str(series).strip():
        full_name = f"[{str(series).strip()}] {full_name}"

    pack_qty, base_unit = _parse_unit_cell(get('unit'))

    barcode = get('barcode')
    barcode_str = str(barcode).strip() if barcode else ''
    barcode_box = get('barcode_box')
    barcode_box_str = str(barcode_box).strip() if barcode_box else ''

    unit_price = _to_decimal(get('unit_price'))
    box_price = _to_decimal(get('box_price'))

    # Derive missing values
    if unit_price is None and box_price is not None and pack_qty:
        unit_price = box_price / Decimal(pack_qty)
    if unit_price is None:
        unit_price = Decimal(0)
    if pack_qty > 1 and box_price is None and unit_price:
        box_price = unit_price * Decimal(pack_qty)

    # Build packagings
    packagings = [{
        'name': base_unit or '單個',
        'quantity': 1,
        'price': float(unit_price.quantize(Decimal('0.001'))),
        'cost': float(unit_price.quantize(Decimal('0.001'))),
        'barcode': barcode_str,
        'is_default': True,
    }]
    if pack_qty > 1 and box_price and box_price > 0:
        packagings.append({
            'name': f'整箱({pack_qty})',
            'quantity': pack_qty,
            'price': float(box_price.quantize(Decimal('0.001'))),
            'cost': float(box_price.quantize(Decimal('0.001'))),
            'barcode': barcode_box_str,
            'is_default': False,
        })

    return {
        'sku': sku,
        'name': full_name,
        'barcode': barcode_str,
        'base_unit': base_unit or '個',
        'safety_stock': 0,
        'packagings': packagings,
    }
